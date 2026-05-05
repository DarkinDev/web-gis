"""
Frontend Views - User-facing views
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
import json
from bus.models import BusRoute, BusStop, RouteStop
from django.core.paginator import Paginator
from .forms import UserUpdateForm, ProfileUpdateForm

# Kiểm tra user có quyền staff (nhân viên hoặc admin)
is_staff = user_passes_test(lambda u: u.is_active and u.is_staff, login_url='/accounts/login/')


@login_required
def profile_view(request):
    """Trang hồ sơ cá nhân: Sửa tên, email, sđt"""
    if request.method == 'POST':
        u_form = UserUpdateForm(request.POST, instance=request.user)
        p_form = ProfileUpdateForm(request.POST, instance=request.user.profile)
        if u_form.is_valid() and p_form.is_valid():
            u_form.save()
            p_form.save()
            messages.success(request, 'Hồ sơ của bạn đã được cập nhật!')
            return redirect('profile')
    else:
        u_form = UserUpdateForm(instance=request.user)
        p_form = ProfileUpdateForm(instance=request.user.profile)

    context = {
        'u_form': u_form,
        'p_form': p_form
    }
    return render(request, 'frontend/profile.html', context)


@login_required
def home_view(request):
    """Home page with full-screen map"""
    routes = BusRoute.objects.filter(is_active=True).order_by('route_number')
    context = {
        'routes': routes,
        'total_routes': routes.count(),
        'total_stops': BusStop.objects.filter(is_active=True).count(),
    }
    return render(request, 'frontend/home.html', context)


@login_required
def routes_list_view(request):
    """List all bus routes with pagination"""
    all_routes = BusRoute.objects.filter(is_active=True).order_by('route_number')
    paginator  = Paginator(all_routes, 10)
    page_num   = request.GET.get('page')
    routes     = paginator.get_page(page_num)
    context    = {'routes': routes}
    return render(request, 'frontend/routes.html', context)


@login_required
def route_detail_view(request, pk):
    """Detail view for a specific route"""
    route = get_object_or_404(BusRoute, pk=pk)
    stops = route.routestop_set.order_by('order').select_related('stop')
    context = {'route': route, 'stops': stops}
    return render(request, 'frontend/route_detail.html', context)


@login_required
def stops_list_view(request):
    """List all bus stops with pagination"""
    all_stops = BusStop.objects.filter(is_active=True).order_by('name')
    paginator = Paginator(all_stops, 10)
    page_num  = request.GET.get('page')
    stops     = paginator.get_page(page_num)
    context   = {'stops': stops}
    return render(request, 'frontend/stops.html', context)


@is_staff
def management_view(request):
    """Management dashboard for CRUD operations"""
    all_routes = BusRoute.all_objects.all().order_by('route_number')
    all_stops  = BusStop.all_objects.all().order_by('name')

    # Phân trang Tuyến xe (10 bản ghi/trang)
    route_paginator = Paginator(all_routes, 10)
    route_page_num  = request.GET.get('route_page')
    routes          = route_paginator.get_page(route_page_num)

    # Phân trang Trạm dừng (10 bản ghi/trang)
    stop_paginator = Paginator(all_stops, 10)
    stop_page_num  = request.GET.get('stop_page')
    stops          = stop_paginator.get_page(stop_page_num)

    context = {
        'routes': routes,
        'stops': stops,
        'total_routes': all_routes.count(),
        'total_stops': all_stops.count(),
    }
    return render(request, 'frontend/management.html', context)


# ─── CRUD API endpoints ────────────────────────────────────────────────────────

@csrf_exempt
@require_http_methods(["POST"])
def api_create_route(request):
    try:
        data = json.loads(request.body)
        route = BusRoute.objects.create(
            route_number=data['route_number'],
            name=data['name'],
            start_point=data.get('start_point', ''),
            end_point=data.get('end_point', ''),
            description=data.get('description', ''),
            operating_hours=data.get('operating_hours', '05:00 - 22:00'),
            frequency=data.get('frequency', '10-15 phút/chuyến'),
            ticket_price=data.get('ticket_price', 7000),
            color=data.get('color', '#3388ff'),
            is_active=data.get('is_active', True),
        )
        return JsonResponse({'success': True, 'id': route.id, 'message': f'Đã tạo tuyến {route.route_number}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_update_route(request, pk):
    try:
        route = get_object_or_404(BusRoute, pk=pk)
        data  = json.loads(request.body)
        for field in ['route_number', 'name', 'start_point', 'end_point',
                      'description', 'operating_hours', 'frequency',
                      'ticket_price', 'color', 'is_active']:
            if field in data:
                setattr(route, field, data[field])
        route.save()
        return JsonResponse({'success': True, 'message': f'Đã cập nhật tuyến {route.route_number}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_delete_route(request, pk):
    try:
        route = get_object_or_404(BusRoute, pk=pk)
        num   = route.route_number
        route.soft_delete()
        return JsonResponse({'success': True, 'message': f'Đã xóa tuyến {num}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_restore_route(request, pk):
    try:
        route = get_object_or_404(BusRoute.all_objects, pk=pk)
        num   = route.route_number
        route.restore()
        return JsonResponse({'success': True, 'message': f'Đã khôi phục tuyến {num}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_create_stop(request):
    try:
        from django.contrib.gis.geos import Point
        data  = json.loads(request.body)
        lat   = float(data['latitude'])
        lng   = float(data['longitude'])
        stop  = BusStop.objects.create(
            name=data['name'],
            code=data.get('code', None) or None,
            address=data.get('address', ''),
            location=Point(lng, lat, srid=4326),
            has_shelter=data.get('has_shelter', False),
            has_bench=data.get('has_bench', False),
            is_active=data.get('is_active', True),
        )
        return JsonResponse({'success': True, 'id': stop.id, 'message': f'Đã tạo trạm "{stop.name}"'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_update_stop(request, pk):
    try:
        from django.contrib.gis.geos import Point
        stop  = get_object_or_404(BusStop, pk=pk)
        data  = json.loads(request.body)
        for field in ['name', 'code', 'address', 'has_shelter', 'has_bench', 'is_active']:
            if field in data:
                setattr(stop, field, data[field])
        if 'latitude' in data and 'longitude' in data:
            stop.location = Point(float(data['longitude']), float(data['latitude']), srid=4326)
        stop.save()
        return JsonResponse({'success': True, 'message': f'Đã cập nhật trạm "{stop.name}"'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_delete_stop(request, pk):
    try:
        stop = get_object_or_404(BusStop, pk=pk)
        name = stop.name
        stop.soft_delete()
        return JsonResponse({'success': True, 'message': f'Đã xóa trạm "{name}"'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_restore_stop(request, pk):
    try:
        stop = get_object_or_404(BusStop.all_objects, pk=pk)
        name = stop.name
        stop.restore()
        return JsonResponse({'success': True, 'message': f'Đã khôi phục trạm "{name}"'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


def api_get_route(request, pk):
    try:
        route = get_object_or_404(BusRoute, pk=pk)
        return JsonResponse({
            'id': route.id, 'route_number': route.route_number, 'name': route.name,
            'start_point': route.start_point, 'end_point': route.end_point,
            'description': route.description, 'operating_hours': route.operating_hours,
            'frequency': route.frequency, 'ticket_price': str(route.ticket_price),
            'color': route.color, 'is_active': route.is_active,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def api_get_stop(request, pk):
    try:
        stop = get_object_or_404(BusStop, pk=pk)
        return JsonResponse({
            'id': stop.id, 'name': stop.name, 'code': stop.code or '',
            'address': stop.address, 'latitude': stop.latitude,
            'longitude': stop.longitude, 'has_shelter': stop.has_shelter,
            'has_bench': stop.has_bench, 'is_active': stop.is_active,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.gis.geos import Point
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from django.core.paginator import Paginator
from django.contrib.auth.models import User
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from bus.models import BusRoute, BusStop, RouteStop

def is_staff(user):
    return user.is_active and user.is_staff

@user_passes_test(is_staff, login_url='/accounts/login/')
def management_view(request):
    routes = BusRoute.all_objects.all().order_by('route_number')
    stops = BusStop.all_objects.all().order_by('name')
    context = {
        'routes': routes,
        'stops': stops,
        'total_routes': BusRoute.objects.count(),
        'total_stops': BusStop.objects.count(),
    }
    return render(request, 'frontend/management.html', context)

@user_passes_test(lambda u: u.is_active and u.is_superuser, login_url='/accounts/login/')
def user_management_view(request):
    users = User.objects.all().order_by('-date_joined')
    total_users = users.count()
    active_users = users.filter(is_active=True).count()
    staff_users = users.filter(is_staff=True).count()
    last_week = timezone.now() - timedelta(days=7)
    new_users_last_week = users.filter(date_joined__gte=last_week).count()
    
    search_query = request.GET.get('q', '').strip()
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
        
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'users': page_obj,
        'total_users': total_users,
        'active_users': active_users,
        'staff_users': staff_users,
        'new_users_last_week': new_users_last_week,
        'search_query': search_query,
    }
    return render(request, 'frontend/user_management.html', context)

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(lambda u: u.is_active and u.is_superuser)
def api_toggle_user_status(request, user_id):
    try:
        if request.user.id == user_id:
            return JsonResponse({'success': False, 'message': 'Không thể khóa tài khoản của chính mình!'})
        user = User.objects.get(id=user_id)
        if user.is_superuser:
            return JsonResponse({'success': False, 'message': 'Không thể khóa tài khoản Admin!'})
        user.is_active = not user.is_active
        user.save()
        status_msg = "Đã mở khóa" if user.is_active else "Đã khóa"
        return JsonResponse({'success': True, 'message': f'{status_msg} tài khoản {user.username}', 'is_active': user.is_active})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy người dùng'})

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(lambda u: u.is_active and u.is_superuser)
def api_toggle_user_staff(request, user_id):
    try:
        if request.user.id == user_id:
            return JsonResponse({'success': False, 'message': 'Không thể tự thay đổi quyền của chính mình!'})
        user = User.objects.get(id=user_id)
        if user.is_superuser:
            return JsonResponse({'success': False, 'message': 'Không thể thay đổi quyền của Admin!'})
        user.is_staff = not user.is_staff
        user.save()
        role_msg = "đã cấp quyền Staff" if user.is_staff else "đã gỡ quyền Staff"
        return JsonResponse({'success': True, 'message': f'Tài khoản {user.username} {role_msg}', 'is_staff': user.is_staff})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy người dùng'})

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(lambda u: u.is_active and u.is_superuser)
def api_update_user_email(request, user_id):
    try:
        data = json.loads(request.body)
        new_email = data.get('email', '').strip()
        user = User.objects.get(id=user_id)
        if user.is_superuser and request.user.id != user_id:
            return JsonResponse({'success': False, 'message': 'Không thể đổi email của Admin khác!'})
        user.email = new_email
        user.save()
        return JsonResponse({'success': True, 'message': f'Đã cập nhật email cho tài khoản {user.username}'})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy người dùng'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(lambda u: u.is_active and u.is_superuser)
def api_reset_user_password(request, user_id):
    try:
        user = User.objects.get(id=user_id)
        if user.is_superuser and request.user.id != user_id:
            return JsonResponse({'success': False, 'message': 'Không thể đổi mật khẩu của Admin khác!'})
        user.set_password('123456')
        user.save()
        return JsonResponse({'success': True, 'message': f'Đã reset mật khẩu tài khoản {user.username} thành: 123456'})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy người dùng'})

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(lambda u: u.is_active and u.is_superuser)
def api_delete_user(request, user_id):
    try:
        if request.user.id == user_id:
            return JsonResponse({'success': False, 'message': 'Không thể tự xóa tài khoản của chính mình!'})
        user = User.objects.get(id=user_id)
        if user.is_superuser:
            return JsonResponse({'success': False, 'message': 'Không thể xóa tài khoản Admin!'})
        username = user.username
        user.delete()
        return JsonResponse({'success': True, 'message': f'Đã xóa vĩnh viễn tài khoản {username}'})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không tìm thấy người dùng'})

@csrf_exempt
@require_http_methods(["GET"])
def api_get_route_stops(request, pk):
    try:
        route = get_object_or_404(BusRoute, pk=pk)
        route_stops = RouteStop.objects.filter(route=route).order_by('order').select_related('stop')
        data = []
        for rs in route_stops:
            data.append({
                'id': rs.id,
                'stop_id': rs.stop.id,
                'stop_name': rs.stop.name,
                'stop_code': rs.stop.code,
                'stop_address': rs.stop.address,
                'order': rs.order,
                'direction': rs.direction,
                'distance_from_start': float(rs.distance_from_start) if rs.distance_from_start else 0,
                'latitude': float(rs.stop.latitude) if rs.stop.latitude else None,
                'longitude': float(rs.stop.longitude) if rs.stop.longitude else None,
            })
        return JsonResponse({'success': True, 'stops': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(is_staff)
def api_add_route_stop(request, pk):
    try:
        data = json.loads(request.body)
        stop_id = data.get('stop_id')
        order = data.get('order', 1)
        direction = data.get('direction', 'outbound')
        
        route = get_object_or_404(BusRoute, pk=pk)
        stop = get_object_or_404(BusStop, pk=stop_id)
        
        rs = RouteStop.objects.create(
            route=route, stop=stop, order=order, direction=direction
        )
        return JsonResponse({'success': True, 'message': 'Đã thêm trạm vào tuyến', 'id': rs.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(is_staff)
def api_remove_route_stop(request, pk, rs_id):
    try:
        rs = get_object_or_404(RouteStop, pk=rs_id, route_id=pk)
        rs.delete()
        return JsonResponse({'success': True, 'message': 'Đã xóa trạm khỏi tuyến'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@csrf_exempt
@require_http_methods(["POST"])
@user_passes_test(is_staff)
def api_reorder_route_stops(request, pk):
    try:
        data = json.loads(request.body)
        order_data = data.get('orders', [])
        
        for item in order_data:
            rs_id = item.get('id')
            new_order = item.get('order')
            RouteStop.objects.filter(id=rs_id, route_id=pk).update(order=new_order)
            
        return JsonResponse({'success': True, 'message': 'Đã cập nhật thứ tự trạm'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@user_passes_test(is_staff)
def export_stops_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trạm Dừng"
    headers = ["ID", "Tên trạm", "Mã trạm", "Địa chỉ", "Vĩ độ", "Kinh độ", "Có nhà chờ", "Trạng thái"]
    ws.append(headers)
    for s in BusStop.all_objects.all().order_by('name'):
        ws.append([s.id, s.name, s.code, s.address, s.latitude, s.longitude, "Có" if s.has_shelter else "Không", "Hoạt động" if s.is_active else "Đã xóa"])
    res = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    res['Content-Disposition'] = 'attachment; filename="tram_dung.xlsx"'
    wb.save(res)
    return res

@user_passes_test(is_staff)
def export_routes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tuyến Xe"
    headers = ["Số tuyến", "Tên tuyến", "Điểm đầu", "Điểm cuối", "Giờ HĐ", "Trạng thái"]
    ws.append(headers)
    for r in BusRoute.all_objects.all().order_by('route_number'):
        ws.append([r.route_number, r.name, r.start_point, r.end_point, r.operating_hours, "Hoạt động" if r.is_active else "Đã xóa"])
    res = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    res['Content-Disposition'] = 'attachment; filename="tuyen_xe.xlsx"'
    wb.save(res)
    return res

@user_passes_test(is_staff)
def export_route_stops_excel(request, pk):
    route = get_object_or_404(BusRoute, pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Trạm tuyến {route.route_number}"
    headers = ["Thứ tự", "Chiều", "Tên trạm", "Mã trạm", "Địa chỉ"]
    ws.append(headers)
    for rs in RouteStop.objects.filter(route=route).order_by('order').select_related('stop'):
        ws.append([rs.order, "Lượt đi" if rs.direction == 'outbound' else "Lượt về", rs.stop.name, rs.stop.code, rs.stop.address])
    res = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    res['Content-Disposition'] = f'attachment; filename="tram_tuyen_{route.route_number}.xlsx"'
    wb.save(res)
    return res

@user_passes_test(lambda u: u.is_active and u.is_superuser)
def export_users_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách người dùng"
    headers = ["ID", "Tên đăng nhập", "Họ Tên", "Email", "Vai trò", "Trạng thái", "Ngày tham gia"]
    ws.append(headers)
    users = User.objects.all().order_by('-date_joined')
    for u in users:
        role = "Admin" if u.is_superuser else ("Staff" if u.is_staff else "User")
        status = "Đang hoạt động" if u.is_active else "Bị khóa"
        date_joined = u.date_joined.strftime("%d/%m/%Y %H:%M") if u.date_joined else ""
        ws.append([u.id, u.username, f"{u.first_name} {u.last_name}".strip(), u.email, role, status, date_joined])
    res = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    res['Content-Disposition'] = 'attachment; filename="danh_sach_nguoi_dung.xlsx"'
    wb.save(res)
    return res

@csrf_exempt
@user_passes_test(is_staff)
def import_stops_excel(request):
    return JsonResponse({'success': False, 'error': 'Chưa được cấp quyền upload file import trạm'})

@csrf_exempt
@user_passes_test(is_staff)
def import_routes_excel(request):
    return JsonResponse({'success': False, 'error': 'Chưa được cấp quyền upload file import tuyến'})

def register_view(request):
    from frontend.forms import CustomUserCreationForm
    from django.contrib.auth import login
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

def landing_view(request):
    return render(request, 'landing.html')

@login_required
def stop_detail_view(request, pk):
    stop = get_object_or_404(BusStop, pk=pk)
    routes = BusRoute.objects.filter(routestop__stop=stop, is_active=True).distinct()
    context = {'stop': stop, 'routes': routes}
    return render(request, 'frontend/stop_detail.html', context)

